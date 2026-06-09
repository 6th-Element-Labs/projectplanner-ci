"""On-demand exports from LIVE task state (not static files): Excel (.xlsx) and
Microsoft Project (MSPDI .xml). Ported from build_plan_artifacts.py to read the
store's current rows so exports always reflect edits."""
import datetime
import io
import re
import xml.sax.saxutils as sx
from typing import Any, Dict, List, Tuple

HOURS_PER_DAY = 8


def _ordered(payload: Dict[str, Any]) -> List[Tuple[str, str, Dict[str, Any]]]:
    rows = []
    for w in payload.get("workstreams", []):
        for t in w["tasks"]:
            rows.append((w["workstream_id"], w["name"], t))
    rows.sort(key=lambda x: (x[0], x[2].get("start_day") or 0))
    return rows


PHASE_ORDER = ["Kickoff", "Bootstrap", "Build", "Cutover", "Operate"]


def _all_tasks(payload):
    return [t for w in payload.get("workstreams", []) for t in w.get("tasks", [])]


def _window(tasks):
    starts = sorted(t.get("start_date") for t in tasks if t.get("start_date"))
    fins = sorted(t.get("finish_date") for t in tasks if t.get("finish_date"))
    return (starts[0] if starts else "—"), (fins[-1] if fins else "—")


_DETAIL_HEADERS = ["Workstream", "Task ID", "Task", "Owner Org", "Owner", "Assignee", "Phase",
                   "Status", "Start", "Finish", "Duration (d)", "Effort (d)", "Depends On",
                   "Risk", "Blocking", "Description"]
_DETAIL_WIDTHS = [12, 10, 44, 14, 20, 16, 11, 13, 12, 12, 11, 10, 18, 9, 9, 70]


def _write_task_sheet(sheet, rows):
    """Write a formatted task table. rows = list of (wsid, wsname, task)."""
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    navy, hdr = PatternFill("solid", fgColor="1E3A5F"), Font(bold=True, color="FFFFFF")
    sheet.append(_DETAIL_HEADERS)
    for c in range(1, len(_DETAIL_HEADERS) + 1):
        cell = sheet.cell(row=1, column=c)
        cell.font = hdr
        cell.fill = navy
        cell.alignment = Alignment(vertical="center")

    def _date(v):
        try:
            return datetime.date.fromisoformat(v) if v else None
        except Exception:
            return None

    for wsid, _wsname, t in rows:
        sheet.append([wsid, t["task_id"], t.get("title"), t.get("owner_org"),
                      t.get("owner_person_or_role"), t.get("assignee"), t.get("phase"),
                      t.get("status"), _date(t.get("start_date")), _date(t.get("finish_date")),
                      t.get("duration_days"), t.get("effort_days"),
                      ", ".join(t.get("depends_on", [])), t.get("risk_level"),
                      "Yes" if t.get("is_blocking") else "", t.get("description")])
    for r in range(2, sheet.max_row + 1):
        sheet.cell(row=r, column=9).number_format = "yyyy-mm-dd"
        sheet.cell(row=r, column=10).number_format = "yyyy-mm-dd"
    for i, wdt in enumerate(_DETAIL_WIDTHS, 1):
        sheet.column_dimensions[get_column_letter(i)].width = wdt
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:{get_column_letter(len(_DETAIL_HEADERS))}{sheet.max_row}"


def export_xlsx(payload: Dict[str, Any]) -> bytes:
    """Two sheets: Summary (the exec-summary view) first, then Details (the task table).
    Both reflect whatever filter produced `payload`."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    NAVY = PatternFill("solid", fgColor="1E3A5F")
    HDR = Font(bold=True, color="FFFFFF")
    BOLD = Font(bold=True)
    WRAP = Alignment(wrap_text=True, vertical="top")

    tasks = _all_tasks(payload)
    wb = Workbook()

    # ---------- Sheet 1: Summary ----------
    s = wb.active
    s.title = "Summary"
    s.column_dimensions["A"].width = 26
    for col in "BCDE":
        s.column_dimensions[col].width = 22

    def line(*vals, bold=False, head=False):
        s.append(list(vals))
        r = s.max_row
        for c in range(1, len(vals) + 1):
            cell = s.cell(row=r, column=c)
            if head:
                cell.font = HDR
                cell.fill = NAVY
            elif bold:
                cell.font = BOLD
        return r

    def prose(text):
        s.append([text])
        r = s.max_row
        s.cell(row=r, column=1).alignment = WRAP
        s.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
        s.row_dimensions[r].height = max(30, min(240, (len(text) // 95 + 1) * 15))

    r = line(payload.get("project") or "Project plan")
    s.cell(row=r, column=1).font = Font(bold=True, size=14)
    pstart, pend = _window(tasks)
    line(f"Kickoff: {payload.get('schedule_start') or pstart}    ·    Finish: {pend}")
    if payload.get("schedule_note"):
        prose(payload["schedule_note"])
    line("")

    line("Key numbers", bold=True)
    kpis = [("Workstreams", len(payload.get("workstreams", []))),
            ("Tasks", len(tasks)),
            ("Effort (person-days)", round(sum(t.get("effort_days") or 0 for t in tasks), 1)),
            ("Milestones", len(payload.get("milestones") or [])),
            ("Blocking tasks", sum(1 for t in tasks if t.get("is_blocking"))),
            ("Target finish", pend)]
    for k, v in kpis:
        line(k, v, bold=True)
    line("")

    if payload.get("executive_summary"):
        line("Executive summary", bold=True)
        for para in [p for p in payload["executive_summary"].split("\n") if p.strip()]:
            prose(para)
        line("")
    if payload.get("timeline_note"):
        line("Timeline", bold=True)
        prose(payload["timeline_note"])
        line("")

    line("Phases", bold=True)
    line("Phase", "Tasks", "Effort (d)", "Window", head=True)
    for ph in PHASE_ORDER:
        pt = [t for t in tasks if t.get("phase") == ph]
        if pt:
            a, b = _window(pt)
            line(ph, len(pt), round(sum(t.get("effort_days") or 0 for t in pt), 1), f"{a} → {b}")
    line("")

    if payload.get("milestones"):
        line("Milestones", bold=True)
        line("Milestone", "Target", "Gate criteria", head=True)
        for m in payload["milestones"]:
            rr = line(m.get("name"), m.get("target_week"), m.get("gate_criteria"))
            s.cell(row=rr, column=3).alignment = WRAP
        line("")

    line("Workstreams", bold=True)
    line("ID", "Name", "Tasks", "Effort (d)", "Window", head=True)
    for w in payload.get("workstreams", []):
        a, b = _window(w.get("tasks", []))
        line(w["workstream_id"], w["name"], len(w.get("tasks", [])),
             round(sum(t.get("effort_days") or 0 for t in w.get("tasks", [])), 1), f"{a} → {b}")

    # ---------- Sheet 2: Details (all tasks) ----------
    _write_task_sheet(wb.create_sheet("Details"), _ordered(payload))

    # ---------- One sheet per workstream ----------
    used = {"Summary", "Details"}
    for w in payload.get("workstreams", []):
        name = (re.sub(r"[:\\/?*\[\]]", "_", str(w["workstream_id"]))[:31] or "WS")
        base, n = name, 2
        while name in used:
            name = f"{base[:28]}_{n}"
            n += 1
        used.add(name)
        rows = [(w["workstream_id"], w["name"], t) for t in sorted(w.get("tasks", []), key=lambda x: x.get("start_day") or 0)]
        _write_task_sheet(wb.create_sheet(name), rows)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_mspdi(payload: Dict[str, Any]) -> str:
    def esc(s):
        return sx.escape(str(s if s is not None else ""))

    def dt(d, hh="08:00:00"):
        return f"{d}T{hh}" if d else ""

    start = payload.get("schedule_start") or datetime.date.today().isoformat()
    lines = ['<?xml version="1.0" encoding="UTF-8" standalone="yes"?>',
             '<Project xmlns="http://schemas.microsoft.com/project">',
             '<Name>Project Maxwell.xml</Name>',
             f'<Title>{esc(payload.get("project"))}</Title>',
             f'<StartDate>{dt(start)}</StartDate>',
             '<CalendarUID>1</CalendarUID><ScheduleFromStart>1</ScheduleFromStart>',
             '<DefaultStartTime>08:00:00</DefaultStartTime><DefaultFinishTime>17:00:00</DefaultFinishTime>',
             '<MinutesPerDay>480</MinutesPerDay><MinutesPerWeek>2400</MinutesPerWeek>',
             '<DaysPerMonth>20</DaysPerMonth><DurationFormat>7</DurationFormat>',
             '<Calendars><Calendar><UID>1</UID><Name>Standard</Name><IsBaseCalendar>1</IsBaseCalendar><WeekDays>']
    for day in range(1, 8):
        if 2 <= day <= 6:
            lines.append(f'<WeekDay><DayType>{day}</DayType><DayWorking>1</DayWorking><WorkingTimes>'
                         '<WorkingTime><FromTime>08:00:00</FromTime><ToTime>12:00:00</ToTime></WorkingTime>'
                         '<WorkingTime><FromTime>13:00:00</FromTime><ToTime>17:00:00</ToTime></WorkingTime>'
                         '</WorkingTimes></WeekDay>')
        else:
            lines.append(f'<WeekDay><DayType>{day}</DayType><DayWorking>0</DayWorking></WeekDay>')
    lines.append('</WeekDays></Calendar></Calendars><Tasks>')
    lines.append(f'<Task><UID>0</UID><ID>0</ID><Name>{esc(payload.get("project"))}</Name>'
                 '<OutlineLevel>0</OutlineLevel><Summary>1</Summary></Task>')
    uid, tid, uid_by_task = 1, 1, {}
    for w in payload.get("workstreams", []):
        wt = sorted(w["tasks"], key=lambda t: t.get("start_day") or 0)
        if not wt:
            continue
        wstart = min(t.get("start_date") or start for t in wt)
        wfin = max(t.get("finish_date") or start for t in wt)
        lines.append(f'<Task><UID>{uid}</UID><ID>{tid}</ID>'
                     f'<Name>{esc(w["workstream_id"] + " — " + w["name"])}</Name>'
                     f'<OutlineLevel>1</OutlineLevel><Summary>1</Summary>'
                     f'<Start>{dt(wstart)}</Start><Finish>{dt(wfin, "17:00:00")}</Finish></Task>')
        uid += 1; tid += 1
        for t in wt:
            uid_by_task[t["task_id"]] = uid
            dur = t.get("duration_days") or 1
            lines.append(
                f'<Task><UID>{uid}</UID><ID>{tid}</ID><Name>{esc(t["task_id"] + " " + (t.get("title") or ""))}</Name>'
                f'<OutlineLevel>2</OutlineLevel><Manual>1</Manual>'
                f'<Start>{dt(t.get("start_date") or start)}</Start>'
                f'<Finish>{dt(t.get("finish_date") or start, "17:00:00")}</Finish>'
                f'<Duration>PT{dur * HOURS_PER_DAY}H0M0S</Duration><DurationFormat>7</DurationFormat>'
                f'<Notes>{esc(t.get("description"))}</Notes>'
                + "".join(f'<PredecessorLink><PredecessorUID>{uid_by_task[d]}</PredecessorUID><Type>1</Type></PredecessorLink>'
                          for d in t.get("depends_on", []) if d in uid_by_task)
                + '</Task>')
            uid += 1; tid += 1
    lines.append('</Tasks></Project>')
    return "\n".join(lines)
