#!/usr/bin/env python3
"""Build HMI + export artifacts for the Project Maxwell plan.

Source of truth: docs/customers/teep-barnett/project-plan.json (full plan).
This script computes a schedule (dependency-aware forward pass, floored by each
phase's start week) and emits, into the HMI static dir:

  - data/teep-project-plan.json   slim + dated board JSON (what the page renders)
  - data/teep-project-plan.xlsx   Excel workbook (opens in Excel; importable to MS Project)
  - data/teep-project-plan.xml    MSPDI (Microsoft Project XML — opens natively)

Scheduling lives HERE (server-side data prep), never in the browser JS.

Usage: python3 docs/customers/teep-barnett/build_plan_artifacts.py [YYYY-MM-DD kickoff]
The kickoff date is a placeholder (relative-week plan has no fixed start); pass
a real Monday to rebase, e.g. `... 2026-07-06`.
"""
import datetime
import json
import os
import sys
import xml.sax.saxutils as sx

_HERE = os.path.dirname(os.path.abspath(__file__))
SRC = os.path.join(_HERE, "plan-docs", "project-plan.json")  # full plan (source of truth)
HMI = _HERE  # build artifacts (xlsx/xml samples) land in the repo root

PHASE_ORDER = ["Kickoff", "Bootstrap", "Build", "Cutover", "Operate"]
# COMPRESSED schedule (aggressive pilot: full delivery ~45-60 days, most live ~day 30).
# Phase floors in BUSINESS DAYS from kickoff; phases overlap heavily and run in parallel.
PHASE_START_BDAY = {"Kickoff": 0, "Bootstrap": 2, "Build": 4, "Cutover": 12, "Operate": 20}
# The whole timeline is scaled to fit this many BUSINESS days (~56 calendar days) — an
# aggressive, parallel-staffed pilot. Ordering/proportions preserved; total span guaranteed.
TARGET_SPAN_BDAY = 40
BDAYS_PER_WEEK = 5
HOURS_PER_DAY = 8

TASK_FIELDS = ["task_id", "title", "description", "owner_org", "owner_person_or_role",
               "phase", "effort_days", "depends_on", "entry_criteria", "exit_criteria",
               "deliverable", "risk_level", "is_blocking", "status"]


def monday_on_or_after(d):
    while d.weekday() != 0:  # 0 = Monday
        d += datetime.timedelta(days=1)
    return d


def bd_date(start, n):
    """Calendar date n business days after `start` (n=0 -> start). `start` is a weekday."""
    d = start
    rem = n
    while rem > 0:
        d += datetime.timedelta(days=1)
        if d.weekday() < 5:
            rem -= 1
    return d


def schedule(tasks_by_id):
    """Forward pass: start = max(phase floor, max(dep finishes)). Returns offsets in business days."""
    start_off, end_off = {}, {}

    def finish(tid, stack):
        if tid in end_off:
            return end_off[tid]
        t = tasks_by_id[tid]
        floor = PHASE_START_BDAY.get(t["phase"], 0)
        s = floor
        for dep in t.get("depends_on", []):
            if dep in tasks_by_id and dep != tid and dep not in stack:
                s = max(s, finish(dep, stack | {tid}))
        dur = max(1, int(round(t.get("effort_days", 1) or 1)))
        start_off[tid] = s
        end_off[tid] = s + dur
        return end_off[tid]

    for tid in tasks_by_id:
        finish(tid, set())
    return start_off, end_off


def main():
    kickoff = sys.argv[1] if len(sys.argv) > 1 else "2026-07-06"
    start = monday_on_or_after(datetime.date.fromisoformat(kickoff))

    full = json.load(open(SRC))
    tasks_by_id, ws_of = {}, {}
    for w in full["workstreams"]:
        for t in w["tasks"]:
            tasks_by_id[t["task_id"]] = t
            ws_of[t["task_id"]] = w["workstream_id"]

    start_off, end_off = schedule(tasks_by_id)

    # decorate tasks with dates
    max_off = max(end_off.values()) or 1
    factor = min(1.0, TARGET_SPAN_BDAY / max_off)
    for tid, t in tasks_by_id.items():
        so = int(round(start_off[tid] * factor))
        dur = max(1, int(round((end_off[tid] - start_off[tid]) * factor)))
        t["start_date"] = bd_date(start, so).isoformat()
        t["finish_date"] = bd_date(start, so + dur - 1).isoformat()
        t["duration_days"] = dur
        t["start_day"] = so

    os.makedirs(HMI, exist_ok=True)

    # ---- 1) slim + dated board JSON ----
    extra = ["start_date", "finish_date", "duration_days", "start_day"]
    slim = {
        "project": full.get("project"),
        "generated": full.get("generated"),
        "schedule_start": start.isoformat(),
        "schedule_note": f"Aggressive pilot: full delivery in ~45-60 days, most workstreams live by "
                         f"~day 30. Kickoff {start.isoformat()} (Monday); dates computed from dependencies. "
                         "Mark week-1 tasks done on the board as they complete.",
        "owner_orgs": full.get("owner_orgs"),
        "rollups": {k: full["rollups"][k] for k in ["total_workstreams", "total_tasks", "total_effort_days"]},
        "executive_summary": full.get("executive_summary"),
        "timeline_note": (
            "Aggressive pilot target set by the team: full delivery in ~45-60 days with most "
            "workstreams live by ~day 30 — faster than the original 11-15 week 'realistic' "
            "estimate, on the bet that the bootstrap track plus parallel staffing collapse the "
            "critical path. The original go-live gates still apply (Entra/SSO admin consent, "
            "Bedrock model-access, the net-new TaskHub API, the unnamed IFS owner) and must be "
            "front-loaded into week 1 to hold this schedule."
        ),
        "workstreams": [
            {"workstream_id": w["workstream_id"], "name": w["name"],
             "tasks": [{f: t.get(f) for f in TASK_FIELDS + extra} for t in w["tasks"]]}
            for w in full["workstreams"]
        ],
        "critical_path": full.get("critical_path"),
        "milestones": full.get("milestones"),
        "consolidated_risks": full.get("consolidated_risks"),
        "consolidated_decisions": full.get("consolidated_decisions"),
    }
    json.dump(slim, open(os.path.join(_HERE, "seed_plan.json"), "w"),
              separators=(",", ":"), ensure_ascii=False)

    # ---- 2) Excel (.xlsx) ----
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Project Plan"
    headers = ["Workstream", "Task ID", "Task", "Owner Org", "Owner", "Phase",
               "Start", "Finish", "Duration (d)", "Effort (d)", "Depends On",
               "Risk", "Blocking", "Status", "Description"]
    ws.append(headers)
    hdr_fill = PatternFill("solid", fgColor="1E3A5F")
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = hdr_fill
        cell.alignment = Alignment(vertical="center")
    ordered = []
    for w in full["workstreams"]:
        for t in w["tasks"]:
            ordered.append((w["workstream_id"], t))
    ordered.sort(key=lambda x: (x[0], x[1]["start_day"]))
    for wsid, t in ordered:
        ws.append([
            wsid, t["task_id"], t["title"], t["owner_org"], t["owner_person_or_role"], t["phase"],
            datetime.date.fromisoformat(t["start_date"]), datetime.date.fromisoformat(t["finish_date"]),
            t["duration_days"], t.get("effort_days"), ", ".join(t.get("depends_on", [])),
            t["risk_level"], "Yes" if t["is_blocking"] else "", t.get("status", "Not Started"),
            t["description"],
        ])
    for r in range(2, ws.max_row + 1):
        ws.cell(row=r, column=7).number_format = "yyyy-mm-dd"
        ws.cell(row=r, column=8).number_format = "yyyy-mm-dd"
    widths = [12, 10, 46, 14, 22, 11, 12, 12, 11, 10, 18, 9, 9, 12, 70]
    for i, wdt in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = wdt
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"

    notes = wb.create_sheet("Notes")
    notes["A1"] = "Project Maxwell — TEEP Barnett Phase-1 Pilot"
    notes["A1"].font = Font(bold=True, size=14)
    notes["A3"] = slim["schedule_note"]
    notes["A5"] = f"Workstreams: {slim['rollups']['total_workstreams']}  ·  Tasks: {slim['rollups']['total_tasks']}  ·  Effort: {slim['rollups']['total_effort_days']} person-days"
    notes.column_dimensions["A"].width = 120
    wb.save(os.path.join(HMI, "teep-project-plan.xlsx"))

    # ---- 3) MSPDI (Microsoft Project XML) ----
    def esc(s):
        return sx.escape(str(s if s is not None else ""))

    def iso_dt(d, hh="08:00:00"):
        return f"{d}T{hh}"

    lines = ['<?xml version="1.0" encoding="UTF-8" standalone="yes"?>',
             '<Project xmlns="http://schemas.microsoft.com/project">',
             '<Name>Project Maxwell.xml</Name>',
             f'<Title>{esc(full.get("project"))}</Title>',
             f'<StartDate>{iso_dt(start.isoformat())}</StartDate>',
             '<CalendarUID>1</CalendarUID>',
             '<ScheduleFromStart>1</ScheduleFromStart>',
             '<DefaultStartTime>08:00:00</DefaultStartTime>',
             '<DefaultFinishTime>17:00:00</DefaultFinishTime>',
             '<MinutesPerDay>480</MinutesPerDay>',
             '<MinutesPerWeek>2400</MinutesPerWeek>',
             '<DaysPerMonth>20</DaysPerMonth>',
             '<DurationFormat>7</DurationFormat>',
             '<Calendars><Calendar><UID>1</UID><Name>Standard</Name><IsBaseCalendar>1</IsBaseCalendar>',
             '<WeekDays>']
    # Sun(1)/Sat(7) off; Mon-Fri working 08:00-12:00,13:00-17:00
    for day in range(1, 8):
        working = 1 if 2 <= day <= 6 else 0
        if working:
            lines.append(f'<WeekDay><DayType>{day}</DayType><DayWorking>1</DayWorking>'
                         '<WorkingTimes>'
                         '<WorkingTime><FromTime>08:00:00</FromTime><ToTime>12:00:00</ToTime></WorkingTime>'
                         '<WorkingTime><FromTime>13:00:00</FromTime><ToTime>17:00:00</ToTime></WorkingTime>'
                         '</WorkingTimes></WeekDay>')
        else:
            lines.append(f'<WeekDay><DayType>{day}</DayType><DayWorking>0</DayWorking></WeekDay>')
    lines.append('</WeekDays></Calendar></Calendars>')
    lines.append('<Tasks>')

    uid = [0]
    # UID 0 = project summary
    lines.append(f'<Task><UID>0</UID><ID>0</ID><Name>{esc(full.get("project"))}</Name>'
                 '<OutlineLevel>0</OutlineLevel><OutlineNumber>0</OutlineNumber><Summary>1</Summary></Task>')
    uid_by_task = {}
    next_uid = 1
    next_id = 1
    # group by workstream: level-1 summary per workstream, level-2 child tasks
    for w in full["workstreams"]:
        ws_uid = next_uid; next_uid += 1
        ws_id = next_id; next_id += 1
        wtasks = sorted(w["tasks"], key=lambda t: t["start_day"])
        wstart = min(t["start_date"] for t in wtasks)
        wfin = max(t["finish_date"] for t in wtasks)
        lines.append(f'<Task><UID>{ws_uid}</UID><ID>{ws_id}</ID>'
                     f'<Name>{esc(w["workstream_id"] + " — " + w["name"])}</Name>'
                     f'<OutlineLevel>1</OutlineLevel><Summary>1</Summary>'
                     f'<Start>{iso_dt(wstart)}</Start><Finish>{iso_dt(wfin, "17:00:00")}</Finish></Task>')
        for t in wtasks:
            tu = next_uid; next_uid += 1
            ti = next_id; next_id += 1
            uid_by_task[t["task_id"]] = tu
            mins = t["duration_days"] * HOURS_PER_DAY * 60
            lines.append(
                f'<Task><UID>{tu}</UID><ID>{ti}</ID><Name>{esc(t["task_id"] + " " + t["title"])}</Name>'
                f'<OutlineLevel>2</OutlineLevel><Manual>1</Manual>'
                f'<Start>{iso_dt(t["start_date"])}</Start><Finish>{iso_dt(t["finish_date"], "17:00:00")}</Finish>'
                f'<Duration>PT{t["duration_days"] * HOURS_PER_DAY}H0M0S</Duration><DurationFormat>7</DurationFormat>'
                f'<Work>PT{t["duration_days"] * HOURS_PER_DAY}H0M0S</Work>'
                f'<Milestone>0</Milestone><Notes>{esc(t["description"])}</Notes>'
                + "".join(
                    f'<PredecessorLink><PredecessorUID>{uid_by_task[d]}</PredecessorUID><Type>1</Type></PredecessorLink>'
                    for d in t.get("depends_on", []) if d in uid_by_task)
                + '</Task>')
    lines.append('</Tasks></Project>')
    with open(os.path.join(HMI, "teep-project-plan.xml"), "w", encoding="utf-8") as f:
        f.write("\n".join(lines))

    print(f"kickoff (placeholder Monday): {start.isoformat()}")
    print(f"tasks scheduled: {len(tasks_by_id)}")
    span_start = min(t['start_date'] for t in tasks_by_id.values())
    span_fin = max(t['finish_date'] for t in tasks_by_id.values())
    print(f"plan span: {span_start} -> {span_fin}")
    for f in ["seed_plan.json", "teep-project-plan.xlsx", "teep-project-plan.xml"]:
        print(f"  wrote {f}: {os.path.getsize(os.path.join(HMI, f))} bytes")


if __name__ == "__main__":
    main()
